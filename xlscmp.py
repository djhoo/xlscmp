import openpyxl
import openpyxl.utils.cell as xlcell
import configparser
import tkinter as tk
from tkinter import messagebox
import sys

message = f"无"
def compare_sheet(sheet1,sheet2,col1,col2,col3):
    global message
    for row_number, (row1, row2) in enumerate(zip(sheet1.iter_rows(min_row=8), sheet2.iter_rows(min_row=8)), start=2):
    # 获取 A 列、B 列和 O 列的值
        a1, b1, o1 = row1[col1].value, row1[col2].value, row1[col3].value
        a2, b2, o2 = row2[col1].value, row2[col2].value, row2[col3].value
        # 将 B 列和 O 列的空值转换为字符串
        b1 = '' if b1 is None else str(b1)
        o1 = '' if o1 is None else str(o1)
        b2 = '' if b2 is None else str(b2)
        o2 = '' if o2 is None else str(o2)

        # 如果 A 列的值不同，则输出该行的行号
        if a1 != a2:
            message += f"{sheet1.title} 的 料号 的值在第 {row_number-1} 行不同\n"
        else:
            # 如果 B 列和 O 列的值不同，则输出该行的行号
            if b1 != b2:
                #print(f"{sheet1.title} 的 第 {row_number-1} 行的 版本 的值不同")
                message += f"{sheet1.title} 的 第 {row_number-1} 行的 版本 的值不同\n"
            if o1 != o2:
                #print(f"{sheet1.title} 的 第 {row_number-1} 行的 每台用量 的值不同")
                message += f"{sheet1.title} 的 第 {row_number-1} 行的 版本 的值不同\n"

# 打开 xlsx 文件
# 创建 ConfigParser 对象
config = configparser.ConfigParser()

# 读取 INI 文件
with open('config.ini', 'r', encoding='utf-8-sig') as f:
    config.read_file(f)
file1 = config.get('file', 'file1')
file2 = config.get('file', 'file2')
try:
    workbook1 = openpyxl.load_workbook(file1)
    workbook2 = openpyxl.load_workbook(file2)
except:
    messagebox.showinfo("提示", "没有找到文件")
    sys.exit()

# 获取 sheet 对象
sheet1 = workbook1['泰山多功能整机']
sheet2 = workbook2['泰山多功能整机']
message = f""
#O列15，R列18，AQ列43
# 遍历两个工作表的行
compare_sheet(sheet1,sheet2,14,17,42)

sheet1 = workbook1['泰山多功能胶囊 ']
sheet2 = workbook2['泰山多功能胶囊 ']
#O,R,AI ,15 ,18,35
compare_sheet(sheet1,sheet2,14,17,34)

#泰山多功能AIO 
sheet1 = workbook1['泰山多功能AIO ']
sheet2 = workbook2['泰山多功能AIO ']
#O,R,AJ,15,18,36
compare_sheet(sheet1,sheet2,14,17,35)

#泰山单功能整机
#O,R,X,15,18,24
sheet1 = workbook1['泰山单功能整机']
sheet2 = workbook2['泰山单功能整机']
compare_sheet(sheet1,sheet2,14,17,23)

#泰山单功能AIO
#O,R,X
sheet1 = workbook1['泰山单功能AIO']
sheet2 = workbook2['泰山单功能AIO']
compare_sheet(sheet1,sheet2,14,17,23)

#泰山单功能胶囊
#O,R,X
sheet1 = workbook1['泰山单功能胶囊']
sheet2 = workbook2['泰山单功能胶囊']
compare_sheet(sheet1,sheet2,14,17,23)

messagebox.showinfo("提示", message)


#    print(ao_value)
